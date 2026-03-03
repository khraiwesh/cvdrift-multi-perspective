"""
Compare CVDrift, MDD, and Object-Centric drift detection methods.
GT: 37% and 75% of dataset size.  Tolerance: 10% of dataset size.
Produces a combined Size x Noise cross-tabulation.
"""
import os, re, ast, sys
import pandas as pd
import openpyxl
from collections import defaultdict

BASE = os.path.dirname(os.path.abspath(__file__))


# ──────────────────────── helpers ────────────────────────

def extract_size(name):
    m = re.search(r'(\d+)cases', name.lower())
    return int(m.group(1)) if m else None

def extract_noise(name):
    m = re.search(r'noisy_(\d+)%', name)
    return f"{m.group(1)}%" if m else "0%"

def extract_interval(name):
    """Extract time interval from filename, e.g. 'dataset_1000cases_10min_ABD' -> '10min'."""
    m = re.search(r'(\d+)min', name.lower())
    return f"{m.group(1)}min" if m else None

def gt_for(n):
    return [int(round(0.37 * n)), int(round(0.75 * n))]

def tolerance_for(n):
    return int(0.10 * n)

def tp_fp_fn(detected, actual, tol):
    if not actual:
        return 0, len(detected), 0
    if not detected:
        return 0, 0, len(actual)
    matched_a, matched_d = set(), set()
    for d in detected:
        for i, a in enumerate(actual):
            if abs(d - a) <= tol and i not in matched_a:
                matched_a.add(i)
                matched_d.add(d)
                break
    tp = len(matched_a)
    return tp, len(detected) - len(matched_d), len(actual) - len(matched_a)

def prf(tp, fp, fn):
    p = tp / (tp + fp) if (tp + fp) else 0
    r = tp / (tp + fn) if (tp + fn) else 0
    f = 2*p*r / (p+r) if (p+r) else 0
    return p, r, f


# ──────────────────────── loaders ────────────────────────

def load_cvdrift():
    """Load CVDrift results from resultsFinal.cvs"""
    path = os.path.join(BASE, "resultsFinal.cvs")
    df = pd.read_csv(path)
    rows = {}
    for _, r in df.iterrows():
        log = r["Log"]
        det_str = r["Detected Changepoints"]
        if det_str == "ERROR" or pd.isna(det_str):
            det = []
        else:
            try: det = ast.literal_eval(det_str)
            except: det = []
        rows[log] = det
    return rows

def load_mdd():
    """Load MDD results from Others/MDD/results.csv"""
    path = os.path.join(BASE, "Others", "MDD", "results.csv")
    df = pd.read_csv(path)
    rows = {}
    for _, r in df.iterrows():
        log = r["Log"]
        det_str = r["Detected Changepoints"]
        if det_str == "ERROR" or pd.isna(det_str):
            det = []
        else:
            try: det = ast.literal_eval(det_str)
            except: det = []
        rows[log] = det
    return rows

def load_oc():
    """Load Object-Centric results from Others/ex_concept_drift/results.xlsx"""
    path = os.path.join(BASE, "Others", "ex_concept_drift", "results.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = {}
    for r in range(2, ws.max_row + 1):
        log = ws.cell(r, 1).value
        det_str = ws.cell(r, 2).value
        err = ws.cell(r, 6).value
        if err or not det_str or det_str == "ERROR":
            det = []
        else:
            try: det = ast.literal_eval(det_str)
            except: det = []
        rows[log] = det
    return rows


# ──────────────────────── main ───────────────────────────

def main():
    print("Loading results …")
    cvd = load_cvdrift()
    mdd = load_mdd()
    oc  = load_oc()

    # Collect all unique log names
    all_logs = sorted(set(cvd) | set(mdd) | set(oc))

    methods = [("CVDrift", cvd), ("MDD", mdd), ("ObjCentric", oc)]

    # Accumulators:  key = (method, size, noise)  and (method, interval)
    acc = defaultdict(lambda: {"tp": 0, "fp": 0, "fn": 0, "n": 0})
    acc_interval = defaultdict(lambda: {"tp": 0, "fp": 0, "fn": 0, "n": 0})
    overall = defaultdict(lambda: {"tp": 0, "fp": 0, "fn": 0, "n": 0})

    for log in all_logs:
        sz = extract_size(log)
        if sz is None:
            continue
        ns = extract_noise(log)
        iv = extract_interval(log)
        gt = gt_for(sz)
        tol = tolerance_for(sz)

        for mname, mdata in methods:
            det = mdata.get(log, None)
            if det is None:
                continue          # method didn't run this log
            tp, fp, fn = tp_fp_fn(det, gt, tol)
            acc[(mname, sz, ns)]["tp"] += tp
            acc[(mname, sz, ns)]["fp"] += fp
            acc[(mname, sz, ns)]["fn"] += fn
            acc[(mname, sz, ns)]["n"]  += 1
            if iv:
                acc_interval[(mname, iv)]["tp"] += tp
                acc_interval[(mname, iv)]["fp"] += fp
                acc_interval[(mname, iv)]["fn"] += fn
                acc_interval[(mname, iv)]["n"]  += 1
            overall[mname]["tp"] += tp
            overall[mname]["fp"] += fp
            overall[mname]["fn"] += fn
            overall[mname]["n"]  += 1

    sizes  = sorted({k[1] for k in acc})
    noises = sorted({k[2] for k in acc}, key=lambda x: int(x.replace('%', '')))
    intervals = sorted({k[1] for k in acc_interval}, key=lambda x: int(x.replace('min', '')))
    mnames = [m for m, _ in methods]

    # ────────────── Pretty-print table ──────────────
    sep = "-"
    hdr_parts = [f"{'Size':>5s}", f"{'Noise':>6s}"]
    for m in mnames:
        hdr_parts.append(f"{'Files':>3s}")
        hdr_parts.append(f"{'TP':>3s}")
        hdr_parts.append(f"{'FP':>3s}")
        hdr_parts.append(f"{'FN':>3s}")
        hdr_parts.append(f"{'P':>6s}")
        hdr_parts.append(f"{'R':>6s}")
        hdr_parts.append(f"{'F1':>6s}")
    hdr = " | ".join(hdr_parts)

    # Method header line
    mhdr_parts = [f"{'':>5s}", f"{'':>6s}"]
    for m in mnames:
        block = f"{m:^41s}"
        mhdr_parts.append(block)
    mhdr = " | ".join(mhdr_parts)

    print("\n" + "=" * len(hdr))
    print("SIZE x NOISE CROSS-TABULATION  —  CVDrift vs MDD vs Object-Centric")
    print("GT = 37% & 75% of N,  Tolerance = 10% of N")
    print("=" * len(hdr))
    print(mhdr)
    print(hdr)
    print(sep * len(hdr))

    csv_rows = []

    for sz in sizes:
        for ns in noises:
            label_ns = "Clean" if ns == "0%" else ns
            parts = [f"{sz:>5d}", f"{label_ns:>6s}"]
            row_data = {"Size": sz, "Noise": label_ns}
            for m in mnames:
                key = (m, sz, ns)
                s = acc.get(key, {"tp": 0, "fp": 0, "fn": 0, "n": 0})
                if s["n"] == 0:
                    parts += [f"{'–':>3s}"] * 7
                    row_data[f"{m}_Files"] = 0
                    for metric in ["TP","FP","FN","P","R","F1"]:
                        row_data[f"{m}_{metric}"] = ""
                else:
                    p, r, f = prf(s["tp"], s["fp"], s["fn"])
                    parts.append(f"{s['n']:>3d}")
                    parts.append(f"{s['tp']:>3d}")
                    parts.append(f"{s['fp']:>3d}")
                    parts.append(f"{s['fn']:>3d}")
                    parts.append(f"{p:>6.4f}")
                    parts.append(f"{r:>6.4f}")
                    parts.append(f"{f:>6.4f}")
                    row_data[f"{m}_Files"] = s["n"]
                    row_data[f"{m}_TP"] = s["tp"]
                    row_data[f"{m}_FP"] = s["fp"]
                    row_data[f"{m}_FN"] = s["fn"]
                    row_data[f"{m}_P"]  = round(p, 4)
                    row_data[f"{m}_R"]  = round(r, 4)
                    row_data[f"{m}_F1"] = round(f, 4)
            print(" | ".join(parts))
            csv_rows.append(row_data)

    # Per-size subtotals
    print(sep * len(hdr))
    for sz in sizes:
        parts = [f"{sz:>5d}", f"{'ALL':>6s}"]
        row_data = {"Size": sz, "Noise": "ALL"}
        for m in mnames:
            tp_t = sum(acc.get((m, sz, ns), {"tp":0})["tp"] for ns in noises)
            fp_t = sum(acc.get((m, sz, ns), {"fp":0})["fp"] for ns in noises)
            fn_t = sum(acc.get((m, sz, ns), {"fn":0})["fn"] for ns in noises)
            n_t  = sum(acc.get((m, sz, ns), {"n":0})["n"]  for ns in noises)
            if n_t == 0:
                parts += [f"{'–':>3s}"] * 7
            else:
                p, r, f = prf(tp_t, fp_t, fn_t)
                parts += [f"{n_t:>3d}", f"{tp_t:>3d}", f"{fp_t:>3d}", f"{fn_t:>3d}",
                          f"{p:>6.4f}", f"{r:>6.4f}", f"{f:>6.4f}"]
            row_data[f"{m}_Files"] = n_t
            row_data[f"{m}_TP"] = tp_t
            row_data[f"{m}_FP"] = fp_t
            row_data[f"{m}_FN"] = fn_t
            pp, rr, ff = prf(tp_t, fp_t, fn_t) if n_t else (0,0,0)
            row_data[f"{m}_P"]  = round(pp, 4)
            row_data[f"{m}_R"]  = round(rr, 4)
            row_data[f"{m}_F1"] = round(ff, 4)
        print(" | ".join(parts))
        csv_rows.append(row_data)

    # Overall
    print(sep * len(hdr))
    parts = [f"{'ALL':>5s}", f"{'ALL':>6s}"]
    row_data = {"Size": "ALL", "Noise": "ALL"}
    for m in mnames:
        s = overall[m]
        p, r, f = prf(s["tp"], s["fp"], s["fn"])
        parts += [f"{s['n']:>3d}", f"{s['tp']:>3d}", f"{s['fp']:>3d}", f"{s['fn']:>3d}",
                  f"{p:>6.4f}", f"{r:>6.4f}", f"{f:>6.4f}"]
        row_data[f"{m}_Files"] = s["n"]
        row_data[f"{m}_TP"]  = s["tp"]
        row_data[f"{m}_FP"]  = s["fp"]
        row_data[f"{m}_FN"]  = s["fn"]
        row_data[f"{m}_P"]   = round(p, 4)
        row_data[f"{m}_R"]   = round(r, 4)
        row_data[f"{m}_F1"]  = round(f, 4)
    print(" | ".join(parts))
    csv_rows.append(row_data)

    # ────────────── TABLE: Per Size ──────────────
    def print_simple_table(title, row_label, keys, agg_fn):
        """Print a compact method-comparison table."""
        shdr = f"{row_label:>8s}"
        for m in mnames:
            shdr += f" | {m+' P':>10s} {m+' R':>10s} {m+' F1':>10s}"
        print("\n" + "=" * len(shdr))
        print(title)
        print("=" * len(shdr))
        print(shdr)
        print("-" * len(shdr))
        for k in keys:
            parts = [f"{str(k):>8s}"]
            for m in mnames:
                tp_t, fp_t, fn_t, n_t = agg_fn(m, k)
                if n_t == 0:
                    parts.append(f"{'–':>10s} {'–':>10s} {'–':>10s}")
                else:
                    p, r, f = prf(tp_t, fp_t, fn_t)
                    parts.append(f"{p:>10.4f} {r:>10.4f} {f:>10.4f}")
            print(" | ".join(parts))
        # Overall row
        parts = [f"{'Overall':>8s}"]
        for m in mnames:
            s = overall[m]
            p, r, f = prf(s["tp"], s["fp"], s["fn"])
            parts.append(f"{p:>10.4f} {r:>10.4f} {f:>10.4f}")
        print("-" * len(shdr))
        print(" | ".join(parts))

    # Per-Size aggregation
    def agg_by_size(m, sz):
        tp_t = sum(acc.get((m, sz, ns), {"tp":0})["tp"] for ns in noises)
        fp_t = sum(acc.get((m, sz, ns), {"fp":0})["fp"] for ns in noises)
        fn_t = sum(acc.get((m, sz, ns), {"fn":0})["fn"] for ns in noises)
        n_t  = sum(acc.get((m, sz, ns), {"n":0})["n"]  for ns in noises)
        return tp_t, fp_t, fn_t, n_t

    size_labels = [str(s) for s in sizes]
    print_simple_table(
        "RESULTS BY DATASET SIZE  —  CVDrift vs MDD vs Object-Centric",
        "Size", size_labels,
        lambda m, k: agg_by_size(m, int(k))
    )

    # Per-Noise aggregation
    def agg_by_noise(m, ns_key):
        ns = "0%" if ns_key == "Clean" else ns_key
        tp_t = sum(acc.get((m, sz, ns), {"tp":0})["tp"] for sz in sizes)
        fp_t = sum(acc.get((m, sz, ns), {"fp":0})["fp"] for sz in sizes)
        fn_t = sum(acc.get((m, sz, ns), {"fn":0})["fn"] for sz in sizes)
        n_t  = sum(acc.get((m, sz, ns), {"n":0})["n"]  for sz in sizes)
        return tp_t, fp_t, fn_t, n_t

    noise_labels = ["Clean" if n == "0%" else n for n in noises]
    print_simple_table(
        "RESULTS BY NOISE LEVEL  —  CVDrift vs MDD vs Object-Centric",
        "Noise", noise_labels,
        lambda m, k: agg_by_noise(m, k)
    )

    # Per-Drift-Amount (time interval) aggregation
    def agg_by_interval(m, iv):
        s = acc_interval.get((m, iv), {"tp":0, "fp":0, "fn":0, "n":0})
        return s["tp"], s["fp"], s["fn"], s["n"]

    print_simple_table(
        "RESULTS BY DRIFT AMOUNT (Time Interval)  —  CVDrift vs MDD vs Object-Centric",
        "Interval", intervals,
        lambda m, k: agg_by_interval(m, k)
    )

    # Save CSV
    out_path = os.path.join(BASE, "comparison_all_methods.csv")
    pd.DataFrame(csv_rows).to_csv(out_path, index=False)
    print(f"\nSaved to {out_path}")


if __name__ == "__main__":
    main()
